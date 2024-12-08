from django.shortcuts import render, get_object_or_404, redirect
from django.core.paginator import Paginator
from django.urls import reverse
from django.http import HttpResponse
from jobscanner.models import Recrutier, Attendee, ScanLog

from io import BytesIO
import qrcode
import openpyxl
from openpyxl.drawing.image import Image
from openpyxl.styles import Alignment
import random

# Create your views here.
PAGE_SIZE = 10

# Helper Functions
def get_hostname(request):
    return f"{request.scheme}://{request.get_host()}"


def qr_generator(host_name: str, freelancer: Attendee):
    url = f"{host_name}/profile/{freelancer.pk}"
    qr = qrcode.make(url)
    buffer = BytesIO()
    qr.save(buffer, format='PNG')
    buffer.seek(0)

    return buffer

def get_login_code(code_list):
    code = random.randint(10000, 50000)
    while code in code_list:
        code = random.randint(10000, 50000)
    
    return code

# Home Page [Login_or_Scan]
def index(request):
    template = "index.html"
    ctx = {}
    if "is_authenticated" in request.session:
        template = "scanner.html"

    if request.method == "POST":
        login_code = request.POST.get("code", "")
        recrutier = Recrutier.objects.filter(code=login_code).first()
        if recrutier:
            request.session['is_authenticated'] = True
            request.session['recrutier_pk'] = recrutier.pk
            return redirect(reverse("home"))
        else:
            ctx['message'] = "Invalid Code!"
    
    return render(request, template, ctx)



# Profile Display [Display profile | Back Home]
def profile(request, pk):
    if "is_authenticated" in request.session:
        freelancer_profile = get_object_or_404(Attendee, pk=pk)
        recrutier = get_object_or_404(Recrutier, pk=request.session['recrutier_pk'])
        scanlog, created = ScanLog.objects.get_or_create(attendee=freelancer_profile, recrutier=recrutier)
        if created:
            freelancer_profile.visits += 1
            recrutier.scanned_counts += 1
            recrutier.save()
            freelancer_profile.save()

        return render(request, "profile.html", {
            "profile" : freelancer_profile,
            'comment' : scanlog.comment
        })
    
    return redirect(reverse("home"))

def comment(request, pk):
   if "is_authenticated" in request.session:
        freelancer_profile = get_object_or_404(Attendee, pk=pk)
        recrutier = get_object_or_404(Recrutier, pk=request.session['recrutier_pk'])
        scan_log, _ = ScanLog.objects.get_or_create(attendee=freelancer_profile, recrutier=recrutier)
        scan_log.comment = request.POST.get("comment", "")
        scan_log.save()

        if _:
            freelancer_profile.visits += 1
            recrutier.scanned_counts += 1
            recrutier.save()
            freelancer_profile.save()

        return redirect(reverse("profile", kwargs={"pk" : pk}))


def scanned(request):
    """
        Filter Scan Logs to each recrutier
        Then split into pages
        Finally return list of comments and freelancers
    """
    if "recrutier_pk" in request.session:
        recrutier = get_object_or_404(Recrutier, pk=request.session['recrutier_pk'])
        scan_log = ScanLog.objects.filter(recrutier=recrutier).order_by("pk")
        paginator = Paginator(scan_log, PAGE_SIZE)
        page_number = request.GET.get('page', "1")
        page_obj = paginator.get_page(page_number)

        return render(request, "scanned_table.html", {
            "page_obj" : page_obj
        })
    return redirect(reverse("home"))


# TODO
def dashboard(request):
    # Display a table with pagination, you can refer to scanned function and scanned_table.html
    # List 10 per page preferable to use PAGE_SIZE
    if request.user.is_authenticated:
        recrutier = Recrutier.objects.all().order_by("name")
        paginator = Paginator(recrutier, PAGE_SIZE)
        page_number = request.GET.get('page', "1")
        page_obj = paginator.get_page(page_number)

        return render(request, "admin_dashboard.html", {
            "page_obj" : page_obj
        })
    return redirect(reverse("home"))

# TODO
def detailed_dashboard(request, pk):
    # Here list all scanned freelancers as you wish as card or in table as you like
    # if you do pagination make sure do not make more than 10 per page and preferable to use PAGE_SIZE
    if request.user.is_authenticated:
        recrutier = get_object_or_404(Recrutier, pk=pk)
        return render(request, "profile_cards.html", {
            "scanlogs" : recrutier.scanned_logs.all()
        })
    return redirect(reverse("home"))

# TODO
def upload_freelancers(request):
    # Take csv file From Zidan [name, email, phone_number, location, track, job_interest, cv_link]
    # Insert in Database using create of Attendee Class
    # then generate csv file [name, email, phone, qr_code]
    # qr_code function that takes host_name and freelancer_obj to build dynamic link, is written for you, it generates buffer to use and create image then embeded it in xlsx
    if request.method == 'POST' and 'file' in request.FILES:
        # 1. Read the uploaded Excel file
        uploaded_file = request.FILES['file']
        wb = openpyxl.load_workbook(uploaded_file)
        sheet = wb.active
        host_name = get_hostname(request)
        
        # 2. Extract data and insert into the Attendee table
        headers = [cell.value for cell in sheet[1]]  # Read headers from the first row
        data_rows = sheet.iter_rows(min_row=2, values_only=True)  # Read data rows
        
        for row in data_rows:
            row_data = dict(zip(headers, row))
            # Map Excel columns to Attendee fields
            try:
                freelancer = Attendee(
                    name=row_data.get("name"),
                    email=row_data.get("email"),
                    phone_number=row_data.get("phone_number"),
                    track=row_data.get("track"),
                    job_interest=row_data.get("job_interest"),
                    cv_url=row_data.get("cv_link"),
                )
                freelancer.save()
            except:
                pass

        # 3. Generate a new Excel file with QR codes
        new_wb = openpyxl.Workbook()
        new_sheet = new_wb.active
        new_sheet.title = "Attendees with QR Codes"

        # Add headers to the new sheet
        headers = ["Name", "Email", "Phone Number", "QR Code"]
        new_sheet.append(headers)

        # Fetch all freelancers and generate QR codes
        freelancers = Attendee.objects.all()
        for freelancer in freelancers:
            # Call the qr_generator function (assume it returns a BytesIO object)
            qr_buffer = qr_generator(host_name, freelancer)
            
            # Write freelancer data into the new sheet
            new_row = [freelancer.name, freelancer.email, freelancer.phone_number]
            new_sheet.append(new_row)
            
            # Insert the QR code image into the fourth column
            qr_image = Image(qr_buffer)
            qr_image.width, qr_image.height = 100, 100
            row_number = new_sheet.max_row  # Get the current row number
            cell_address = f"D{row_number}"  # QR code will be in column D
            new_sheet.row_dimensions[row_number].height = 75
            new_sheet.add_image(qr_image, cell_address)
        
        # Format cells (resize and center align)
        for col in new_sheet.columns:
            max_length = 0
            col_letter = col[0].column_letter  # Get the column letter
            for cell in col:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
                cell.alignment = Alignment(horizontal="center", vertical="center")
            new_sheet.column_dimensions[col_letter].width = max_length + 5  # Add padding

        # 4. Return the new Excel file as a response
        response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        response['Content-Disposition'] = 'attachment; filename="freelancers_with_qrcodes.xlsx"'
        
        output = BytesIO()
        new_wb.save(output)
        output.seek(0)
        response.write(output.getvalue())
        
        return response
    elif request.method == "GET":
        return render(request, "upload_freelancers.html")
    
def upload_recrutiers(request):
    if request.method == 'POST' and 'file' in request.FILES:
        # 1. Read the uploaded Excel file
        uploaded_file = request.FILES['file']
        wb = openpyxl.load_workbook(uploaded_file)
        sheet = wb.active
        codes = []
        
        # 2. Extract data and insert into the Attendee table
        headers = [cell.value for cell in sheet[1]]  # Read headers from the first row
        data_rows = sheet.iter_rows(min_row=2, values_only=True)  # Read data rows
        
        for row in data_rows:
            row_data = dict(zip(headers, row))
            # Map Excel columns to Attendee fields
            recrutier = Recrutier(
                name=row_data.get("name"),
                email=row_data.get("email"),
                members=row_data.get("members"),
                days=row_data.get("days"),
            )
            recrutier.save()

        # 3. Generate a new Excel file with QR codes
        new_wb = openpyxl.Workbook()
        new_sheet = new_wb.active
        new_sheet.title = "Recrutiers with Login Codes"

        # Add headers to the new sheet
        headers = ["Name", "Email", "Login Code"]
        new_sheet.append(headers)

        # Fetch all freelancers and generate QR codes
        recrutiers = Recrutier.objects.all()
        for recrutier in recrutiers:
            # Call the qr_generator function (assume it returns a BytesIO object)
            generated_login_code = get_login_code(codes)
            codes.append(generated_login_code)
            # Write freelancer data into the new sheet
            new_row = [recrutier.name, recrutier.email, generated_login_code]
            new_sheet.append(new_row)

        # 4. Return the new Excel file as a response
        response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        response['Content-Disposition'] = 'attachment; filename="Recrutiers_with_Login_Codes.xlsx"'
        
        output = BytesIO()
        new_wb.save(output)
        output.seek(0)
        response.write(output.getvalue())
        
        return response
    elif request.method == "GET":
        return render(request, "upload_freelancers.html")
    
def download_leads(request):
    if request.user.is_authenticated:
        recrutiers = Recrutier.objects.all()
        new_wb = openpyxl.Workbook()
        for recrutier in recrutiers:
            scan_logs = recrutier.scanned_logs.all()
            new_sheet = new_wb.create_sheet(f"{recrutier.name}")
            headers = ["Name", "Email", "Phone", "ALX Track", "Visits" ]
            new_sheet.append(headers)

            for scan_log in scan_logs:
                attendee = scan_log.attendee
                new_sheet.append([
                    attendee.name,
                    attendee.email,
                    attendee.phone_number,
                    attendee.track,
                    attendee.visits,
                ])

        # 4. Return the new Excel file as a response
        new_wb.remove("Sheet")
        response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        response['Content-Disposition'] = 'attachment; filename="Leads.xlsx"'
        
        output = BytesIO()
        new_wb.save(output)
        output.seek(0)
        response.write(output.getvalue())
        
        return response
    return redirect(reverse("home"))